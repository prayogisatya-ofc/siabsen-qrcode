from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .. import models as db
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.core.paginator import Paginator
import json, locale
import pandas as pd
from django.utils import formats
from datetime import datetime
from django.db.models.functions import ExtractYear, ExtractMonth
from calendar import month_name
from openpyxl import Workbook
from django.utils.translation import gettext as _
from openpyxl.styles import Border, Side


locale.setlocale(locale.LC_TIME, 'id_ID.UTF-8')


@require_GET
@login_required(login_url="/auth/login")
def views(request):
    tahun_absensi = db.AbsensiGuru.objects.annotate(tahun=ExtractYear('tanggal'))
    tahun_unik = tahun_absensi.values_list('tahun', flat=True).distinct()

    bulan_absensi = db.AbsensiGuru.objects.annotate(bulan=ExtractMonth('tanggal'))
    bulan_unik = bulan_absensi.values_list('bulan', flat=True).distinct()

    nama_bulan = month_name[1:]

    select_options = [(bulan, nama_bulan[bulan-1]) for bulan in bulan_unik]

    return render(request, "absensi/absensi.guru.html", {
        'tahun': tahun_unik,
        'bulan': select_options
    })


@require_GET
@login_required(login_url="/auth/login")
def get_data(request):
    page = request.GET.get("page")
    date = request.GET.get("date")
    year = request.GET.get("year")

    data = db.AbsensiGuru.objects.order_by("-id")

    if date:
        data = data.filter(tanggal__month=date)

    if year:
        data = data.filter(tanggal__year=year)

    paginator = Paginator(data, 10)
    current_page = paginator.get_page(page)

    serialized_data = []
    for obj in current_page:
        row = {
            "id": obj.id,
            "tanggal": formats.date_format(obj.tanggal, format='l, d M Y'),
            "absensi": {
                "hadir": obj.detail.all().count(),
                "absen": db.Guru.objects.count() - obj.detail.all().count()
            },
        }
        serialized_data.append(row)
            
    pagination_info = {
        'has_next': current_page.has_next(),
        'has_previous': current_page.has_previous(),
        'current_page_number': current_page.number,
        'total_page': paginator.num_pages,
        'total_data': data.count()
    }
    
    response_data = {
        'data': serialized_data,
        'pagination': pagination_info
    }
        
    return JsonResponse(json.dumps(response_data), safe=False)
    

@require_POST
@login_required(login_url="/auth/login")
def delete_data(request):    
    key = json.loads(request.body).get('key')

    try:
        db.AbsensiGuru.objects.filter(id=key).delete()
        return JsonResponse({'success': 'Absensi berhasil dihapus'})
    except:
        return JsonResponse({'error': "Absensi gagal dihapus!"}, status=400)
    

@require_GET
@login_required(login_url="/auth/login")
def detail_views(request, id):
    return render(request, "absensi/detail.absensi.guru.html")


@require_GET
@login_required(login_url="/auth/login")
def get_detail(request, id):
    data = db.AbsensiGuru.objects.get(id=id)

    detail = []
    for d in db.DetailAbsenGuru.objects.filter(absensi=data).order_by('guru'):
        detail.append({
            'nama': d.guru.nama,
            'foto': d.guru.foto.url if d.guru.foto else None,
            'jam_masuk': d.jam_masuk.strftime('%H:%M') if d.jam_masuk else '-',
            'jam_keluar': d.jam_keluar.strftime('%H:%M') if d.jam_keluar else '-',
        })
    
    row = {
        "id": data.id,
        "tanggal": formats.date_format(data.tanggal, format='l, d M Y'),
        "detail": detail,
    }
    
    return JsonResponse(row)


@require_GET
@login_required(login_url="/auth/login")
def download(request):
    bulan = request.GET.get('month')
    tahun = request.GET.get('year')
    
    absensi_guru = db.AbsensiGuru.objects.filter(
        tanggal__month=bulan,
        tanggal__year=tahun
    )

    # Buat workbook baru
    workbook = Workbook()
    worksheet = workbook.active
    
    # Tambahkan header
    worksheet.append([_("DATA ABSEN GURU SDN 1 AJI JAYA")])
    worksheet.append([_("Bulan"), datetime.strptime(str(bulan), '%m').strftime('%B')])
    worksheet.append([_("Tahun"), tahun])
    worksheet.append([])  # Baris kosong
    
    # Tambahkan header untuk tabel data absensi
    header_row = [_("No."), _("NIP/NUPTK"), _("Nama")]
    # Buat header untuk setiap tanggal dalam bulan tersebut
    for absensi in absensi_guru.all():
        header_row.append(absensi.tanggal.strftime('%Y-%m-%d') + ' Masuk')
        header_row.append(absensi.tanggal.strftime('%Y-%m-%d') + ' Keluar')
    worksheet.append(header_row)
    
    # Tambahkan data absensi
    for i, guru in enumerate(db.Guru.objects.all(), start=1):
        row = [i, guru.kode, guru.nama]
        for absensi in absensi_guru:
            absen = db.DetailAbsenGuru.objects.filter(guru=guru, absensi__tanggal=absensi.tanggal).first()
            if absen:
                row.append(absen.jam_masuk.strftime('%H:%M') if absen.jam_masuk else '-')
                row.append(absen.jam_keluar.strftime('%H:%M') if absen.jam_keluar else '-')
            else:
                row.extend(['-', '-'])
        worksheet.append(row)

    # Menambahkan border pada tabel
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    rows = list(worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column))
    for row in rows:
        for cell in row:
            cell.border = thin_border

    # Buat nama file
    nama_file = f"Data Absen Guru - {datetime.strptime(str(bulan), '%m').strftime('%B')} {tahun}.xlsx"
    
    # Buat respons Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nama_file}"'
    workbook.save(response)
    
    return response